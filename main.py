"""
Arguments:
help - Gives you information
pull - Pulls from the given Google Sheets
pullce - Pulls all objectives from CE.
extract - Pulls the steam app IDs from the Google Sheets and labels them accordingly. Stores them in extract.json
select - Makes a selection based on the extracted data.
    * if extract hasn't been run before, this will fail.
    * this will also print out how many selections it will make per category
output - Dumps a list of Discord strings that need to be sent in the mercantile channel.
    * Formatted like the following:
     # Action
     ## Potentials
     Guardgrave - 3 bp - $4.99 - https://store.steampowered.com/app/2681120/
     Slain: Back from Hell - 3 bp - $12.99 - https://store.steampowered.com/app/369070/
     ## Low Clears
     Blade of Darkness - 2 bp - $14.99 - https://store.steampowered.com/app/1710170/
     ## Uncleareds
     Zero Strain - 10 bp - $9.99 - https://store.steampowered.com/app/1079560/

selectdata - Fetches Steam store prices for all steam entries in selection.json and fills in basePrice/currentPrice.
select+ - Does both select and selectdata.
output - Dumps a list of Discord strings that need to be sent in the mercantile channel.
pull+ - Does both pull and pullce.
"""
import sys
import requests


if len(sys.argv) == 1:
    print("No arguments were provided. Please use 'python main.py help' for help.")
    sys.exit()
if len(sys.argv) > 2:
    print("Incorrect number of arguments passed. Please use 'python main.py help' for help")
    print(f"{sys.argv=}")
    sys.exit()

# config
POTENTIALS_SHEET = "https://docs.google.com/spreadsheets/d/1NeWYzeRi7NDrm9jvJKZgjrB6LLSjKskD3yNO0SYOVpk/edit"
RETRO_SHEET = "https://docs.google.com/spreadsheets/d/1g_7GlGYtz0l4EV_WfWdhslcfYcUZph2-8G4lbHvF214/edit"

# -- number of games ------------
CATEGORY_POTENTIALS_STEAM = 7
CATEGORY_POTENTIALS_RETRO = 3
CATEGORY_LOW_CLEAR_T1 = 3
CATEGORY_LOW_CLEAR_T2 = 2
CATEGORY_LOW_CLEAR_T3 = 2
CATEGORY_LOW_CLEAR_T4 = 1
CATEGORY_LOW_CLEAR_T5PLUS = 1
CATEGORY_UNCLEARED_PO = 2
CATEGORY_UNCLEARED_SO = 1

# -- bounty point worth ---------
BOUNTY_POINTS_POTENTIAL = 3
BOUNTY_POINTS_LOW_CLEAR_T1 = 2
BOUNTY_POINTS_LOW_CLEAR_T2 = 4
BOUNTY_POINTS_LOW_CLEAR_T3 = 8
BOUNTY_POINTS_LOW_CLEAR_T4 = 16
BOUNTY_POINTS_LOW_CLEAR_T5PLUS = 32
BOUNTY_POINTS_UNCLEARED_OBJECTIVE = 10


def pull():
    """
    This function pulls the .xlsx files for both the POTENTIALS_SHEET link and the RETRO_SHEET link
    and stores them as potentials.xlsx and retro.xlsx.
    """
    sheets = [
        (POTENTIALS_SHEET, "potentials.xlsx"),
        (RETRO_SHEET, "retro.xlsx"),
    ]
    for url, filename in sheets:
        sheet_id = url.split("/d/")[1].split("/")[0]
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        print(f"Fetching {filename} (sheet {sheet_id})...")
        response = requests.get(export_url)
        response.raise_for_status()
        with open(filename, "wb") as f:
            f.write(response.content)
        size_kb = len(response.content) / 1024
        print(f"  Saved {filename} ({size_kb:.1f} KB)")
    print("Done. Saved potentials.xlsx and retro.xlsx.")

def extract():
    """
    This function extracts information from potentials.xlsx, retro.xlsx, and uncleareds.json,
    and dumps it all out into totalinfo.json. Each entry in the array should be:
    {
        "platform": "steam" # or "retro"
        "platformId" : "508229" # or a retro id
        "category": "Action" # or "Arcade", "Bullet Hell", "First-Person", "Platformer", "Strategy"
        "type": "Potential" # or "Low Clear" or "Uncleared"
        "lowClearTier": 1 # or 2, 3, 4, 5, null. 5 == "tier 5 or above"
        "gameName": "Celeste" # or whatever
        "ceId": "1f1f23e6-ee09-4029-8973-b6ab79552e17" # or whatever, or null if it's not an uncleared
        "objectiveName": "Thunder and Lightning (UNCLEARED)" # or whatever, or null if it's not an uncleared
        "objectiveCeId": uuidv4
    }
    """
    import json
    import openpyxl

    print("Loading uncleareds.json and cedb_games.json...")
    with open("uncleareds.json") as f:
        uncleareds_raw = json.load(f)
    with open("cedb_games.json") as f:
        cedb_games: dict[str, dict] = json.load(f)
    print(f"  Loaded {len(uncleareds_raw)} uncleareds, {len(cedb_games)} CEDB games.")

    # Transform uncleareds into totalinfo schema
    results = []
    for u in uncleareds_raw:
        if u["gameCeId"] not in cedb_games:
            print(f"  WARNING: gameCeId {u['gameCeId']} ({u['gameName']}) not found in /api/games. Skipping!")
            continue
        game_info = cedb_games[u["gameCeId"]]
        results.append({
            "platform": game_info.get("platform", "steam"),
            "platformId": game_info.get("platformId"),
            "category": game_info.get("category"),
            "type": "Uncleared",
            "lowClearTier": None,
            "gameName": u["gameName"],
            "ceId": u["objectiveCeId"],
            "objectiveName": u["objectiveName"],
            "objectiveCeId": u["objectiveCeId"],
            "objectiveType": u["type"],  # "Primary" or "Secondary"
        })
    print(f"  Transformed {len(results)} uncleareds.")

    def _steam_app_id(url: str | None) -> str | None:
        if url and "/app/" in url:
            try:
                return url.split("/app/")[1].split("/")[0]
            except IndexError:
                pass
        return None

    def _cedb_game_id(url: str | None) -> str | None:
        if url and "cedb.me/game/" in url:
            try:
                return url.rstrip("/").split("/")[-1]
            except IndexError:
                pass
        return None

    # ── potentials.xlsx ──────────────────────────────────────────────────────────
    print("Reading potentials.xlsx...")
    wb = openpyxl.load_workbook("potentials.xlsx", data_only=True)

    STEAM_CATEGORIES = ["Action", "Arcade", "Bullet Hell", "First Person", "Platformer", "Strategy"]

    for sheet_name in STEAM_CATEGORIES:
        ws = wb[sheet_name]
        category = "First-Person" if sheet_name == "First Person" else sheet_name
        rows = list(ws.iter_rows())

        # Strategy has a notes row before the real header; find header by "Game Title"
        header_idx = next(
            i for i, r in enumerate(rows)
            if r[0].value is not None and "Game Title" in str(r[0].value)
        )

        before = len(results)
        for row in rows[header_idx + 1:]:
            name_cell = row[0]
            game_name = name_cell.value
            if game_name is None:
                continue
            platform_id = _steam_app_id(name_cell.hyperlink.target if name_cell.hyperlink else None)
            if platform_id is None:
                print(f"  WARNING: No Steam app ID for '{game_name}' ({category}). Skipping!")
                continue
            results.append({
                "platform": "steam",
                "platformId": platform_id,
                "category": category,
                "type": "Potential",
                "lowClearTier": None,
                "gameName": str(game_name),
                "ceId": None,
                "objectiveName": None,
                "objectiveCeId": None,
            })
        print(f"  {category}: {len(results) - before} potentials")

    # ── Low clears (T1–T5+) ─────────────────────────────────────────────────────
    LOW_CLEAR_TIERS = [
        ("T1 Low Cleared", 1),
        ("T2 Low Cleared", 2),
        ("T3 Low Cleared", 3),
        ("T4 Low Cleared", 4),
        ("T5+ Low Cleared", 5),
    ]
    for sheet_name, tier in LOW_CLEAR_TIERS:
        ws = wb[sheet_name]
        before = len(results)
        for row in ws.iter_rows(min_row=2):
            name_cell = row[0]
            if name_cell.value is None:
                continue
            game_name = str(name_cell.value).replace(" :: CE", "").strip()
            category = row[2].value
            category = category if category != "First Person" else "First-Person"
            ce_game_id = _cedb_game_id(name_cell.hyperlink.target if name_cell.hyperlink else None)
            if not ce_game_id:
                print(f"  WARNING: No CEDB link for low clear '{game_name}' (T{tier}). Skipping!")
                continue
            if ce_game_id not in cedb_games:
                print(f"  WARNING: CEDB game {ce_game_id} ('{game_name}', T{tier}) not found in /api/games. Skipping!")
                continue
            game_info = cedb_games[ce_game_id]
            platform = game_info.get("platform", "steam")
            platform_id = game_info.get("platformId")
            results.append({
                "platform": platform,
                "platformId": platform_id,
                "category": category,
                "type": "Low Clear",
                "lowClearTier": tier,
                "gameName": game_name,
                "ceId": None,
                "objectiveName": None,
                "objectiveCeId": None,
            })
        print(f"  T{tier} Low Clears: {len(results) - before}")

    # ── Steam unlisted check ─────────────────────────────────────────────────
    import time
    import os

    unlisted_path = "unlistedgames.json"
    existing_unlisted: list[dict] = []
    if os.path.exists(unlisted_path):
        with open(unlisted_path) as f:
            existing_unlisted = json.load(f)
    known_unlisted_pids: set[str] = {u["platformId"] for u in existing_unlisted}

    all_steam_ids = {r["platformId"] for r in results if r.get("platform") == "steam" and r.get("platformId")}
    unchecked_ids = list(all_steam_ids - known_unlisted_pids)

    if known_unlisted_pids & all_steam_ids:
        print(f"  Skipping {len(known_unlisted_pids & all_steam_ids)} already-known unlisted game(s).")

    unlisted_ids: set[str] = set(known_unlisted_pids & all_steam_ids)

    if unchecked_ids:
        print(f"Checking {len(unchecked_ids)} Steam app IDs for unlisted status...")
        no_price_ids: list[str] = []
        BATCH = 100
        for i in range(0, len(unchecked_ids), BATCH):
            batch = unchecked_ids[i:i + BATCH]
            resp = requests.get(
                "https://store.steampowered.com/api/appdetails",
                params={"appids": ",".join(batch), "filters": "price_overview", "cc": "us"},
            )
            resp.raise_for_status()
            data = resp.json()
            for app_id, result in data.items():
                if not result.get("success"):
                    no_price_ids.append(app_id)
                    continue
                if not isinstance(result.get("data"), dict):
                    no_price_ids.append(app_id)
            print(f"  Batch {i // BATCH + 1}/{(len(unchecked_ids) + BATCH - 1) // BATCH} done...")
            if i + BATCH < len(unchecked_ids):
                time.sleep(1)

        if no_price_ids:
            print(f"  Second-pass: checking {len(no_price_ids)} no-price entries (free vs unlisted)...")
            for app_id in no_price_ids:
                resp = requests.get(
                    "https://store.steampowered.com/api/appdetails",
                    params={"appids": app_id, "cc": "us"},
                )
                resp.raise_for_status()
                result = resp.json().get(app_id, {})
                if not result.get("success"):
                    unlisted_ids.add(app_id)
                    continue
                app_data = result.get("data")
                if not isinstance(app_data, dict):
                    unlisted_ids.add(app_id)
                    continue
                if not app_data.get("is_free") and not app_data.get("package_groups"):
                    print(f"    {app_data.get('name', app_id)}: unlisted")
                    unlisted_ids.add(app_id)
                time.sleep(0.5)

        new_unlisted = unlisted_ids - known_unlisted_pids
        if new_unlisted:
            pid_to_name = {r["platformId"]: r["gameName"] for r in results if r.get("platformId")}
            for pid in new_unlisted:
                existing_unlisted.append({"platformId": pid, "gameName": pid_to_name.get(pid, "Unknown")})
            with open(unlisted_path, "w") as f:
                json.dump(existing_unlisted, f, indent=2)
            print(f"  {len(new_unlisted)} new unlisted game(s) added to unlistedgames.json.")

    if unlisted_ids:
        before = len(results)
        results = [r for r in results if r.get("platformId") not in unlisted_ids]
        print(f"  Filtered out {before - len(results)} unlisted game(s) from results.")

    with open("totalinfo.json", "w") as f:
        json.dump(results, f, indent=2)

    steam_pots = sum(1 for r in results if r["type"] == "Potential")
    low_clears_total = sum(1 for r in results if r["type"] == "Low Clear")
    uncleareds_total = sum(1 for r in results if r["type"] == "Uncleared")
    print(
        f"Done. Extracted {len(results)} entries total: "
        f"{steam_pots} steam potentials, {uncleareds_total} uncleareds, "
        f"{low_clears_total} low clears. "
        f"Stored in totalinfo.json."
    )

def output():
    """
    Reads selection.json and formats entries into Discord messages (<= 1800 chars each).
    Lines are never split across messages. Saves each message as output01.txt, output02.txt, etc.

    NOTE: technically it's 2000 but i leave in 200 extra characters to put in notes
    """
    import json

    DISCORD_LIMIT = 1800
    CATEGORIES = ["Action", "Arcade", "Bullet Hell", "First-Person", "Platformer", "Strategy"]

    with open("selection.json") as f:
        entries = json.load(f)

    def steam_url(platform_id: str) -> str:
        return f"<https://store.steampowered.com/app/{platform_id}/>"
    def retro_url(platform_id: str) -> str:
        return f"<https://retroachievements.org/game/{platform_id}/>"
    def ce_url(ce_id: str) -> str:
        return f"<https://cedb.me/game/{ce_id}/>"

    def price_str(entry: dict) -> str:
        base = entry.get("basePrice")
        current = entry.get("currentPrice")
        if not current or current == "UNLISTED":
            if base and base != "UNLISTED":
                return base
            if entry['platform'] == 'steam':
                return 'Free!'
            if entry['platform'] == 'retroachievements':
                return 'No price'
            return '?'
        if base and base != current:
            return f"{current} (base {base})"
        return current

    def format_entry(e: dict) -> str:
        price = price_str(e)
        bp = e.get("bountyPoints", "X")
        if e['type'] == 'Uncleared':
            name = f'{e['objectiveType'][0]}O \'{e['objectiveName']}\' - {e['gameName']}'
            url = ce_url(e['ceId'])
            return f"{name} - {bp} bp - {price} - {url}"
        if e.get("platform") == "steam" and e.get("platformId"):
            url = steam_url(e["platformId"])
            return f"{e['gameName']} - {bp} bp - {price} - {url}"
        elif e.get("platform") == "retroachievements" and e.get("platformId"):
            url= retro_url(e["platformId"])
            return f"{e['gameName']} - {bp} bp - {price} - {url}"
        return f"{e['gameName']} - {bp} bp - {price}"

    # Build an ordered list of lines grouped by category and type
    lines: list[str] = []
    for category in CATEGORIES:
        cat = [e for e in entries if e["category"] == category]
        if not cat:
            continue

        lines.append(f"# {category}")

        potentials = [e for e in cat if e["type"] == "Potential"]
        if potentials:
            lines.append("## Potentials")
            for e in potentials:
                lines.append(format_entry(e))

        low_clears = [e for e in cat if e["type"] == "Low Clear"]
        if low_clears:
            lines.append("## Low Clears")
            for e in low_clears:
                lines.append(format_entry(e))

        uncleareds = [e for e in cat if e["type"] == "Uncleared"]
        if uncleareds:
            lines.append("## Uncleareds")
            for e in uncleareds:
                lines.append(format_entry(e))

    # Pack lines into ≤2000-char messages; never split a single line
    messages: list[str] = []
    current_parts: list[str] = []
    current_len = 0

    for line in lines:
        # +1 for the newline that joins parts
        needed = len(line) + (1 if current_parts else 0)
        if current_parts and current_len + needed > DISCORD_LIMIT:
            messages.append("\n".join(current_parts))
            current_parts = [line]
            current_len = len(line)
        else:
            current_parts.append(line)
            current_len += needed

    if current_parts:
        messages.append("\n".join(current_parts))

    for i, msg in enumerate(messages, 1):
        filename = f"output{i:02d}.txt"
        with open(filename, "w") as f:
            f.write(msg)
        print(f"  {filename}: {len(msg)} chars")

    print(f"Done. {len(messages)} message(s) written to output01.txt ... output{len(messages):02d}.txt")


def verify_extract():
    """
    Goes through sheet.json and checks for correctness.
    - No "uncleared" should have a null objectiveName
    - No "uncleared" should have a null ceId
    - No "low clear" should have a null lowClearTier
    - All three of these's inverses should also be true. Every potential should have
      a null ceId and objectiveName.
    """

def select():
    """
    Loads totalinfo.json and randomly selects entries per the configured counts.
    Saves the result to selection.json.

    Per genre category:
      - CATEGORY_POTENTIALS_STEAM potentials
      - CATEGORY_LOW_CLEAR_T1/T2/T3/T4/T5PLUS low clears per tier

    Use the following rules to make RANDOM selections:

    CATEGORY_POTENTIALS_STEAM = 7
    CATEGORY_POTENTIALS_RETRO = 3
    CATEGORY_LOW_CLEAR_T1 = 3
    CATEGORY_LOW_CLEAR_T2 = 2
    CATEGORY_LOW_CLEAR_T3 = 2
    CATEGORY_LOW_CLEAR_T4 = 1
    CATEGORY_LOW_CLEAR_T5PLUS = 1
    
    For each category...
    - CATEGORY_POTENTIALS_STEAM potentials with platform == 'steam'
    - CATEGORY_POTENTIALS_RETRO potentials with platform == 'retro'
      * currently 0 because retros not working
    - CATEGORY_LOW_CLEAR_T1 low clears in tier 1
    - CATEGORY_LOW_CLEAR_T2 low clears in tier 2
    - CATEGORY_LOW_CLEAR_T3 low clears in tier 3
    - CATEGORY_LOW_CLEAR_T4 low clears in tier 4
    - CATEGORY_LOW_CLEAR_T5PLUS low clears in tier 5+
    - CATEGORY_UNCLEARED_PO uncleared POs
    - CATEGORY_UNCLEARED_SO uncleared SOs

    This selection should be relegated to selection.json.
    """
    import json
    import random

    import os

    print("Loading totalinfo.json...")
    with open("totalinfo.json") as f:
        entries = json.load(f)
    print(f"  Loaded {len(entries)} entries.")

    unlisted_pids: set[str] = set()
    if os.path.exists("unlistedgames.json"):
        with open("unlistedgames.json") as f:
            unlisted_pids = {u["platformId"] for u in json.load(f)}
        before = len(entries)
        entries = [e for e in entries if e.get("platformId") not in unlisted_pids]
        print(f"  Excluded {before - len(entries)} unlisted game(s) from pool.")

    CATEGORIES = ["Action", "Arcade", "Bullet Hell", "First-Person", "Platformer", "Strategy"]
    LOW_CLEAR_TIERS = [
        (1, CATEGORY_LOW_CLEAR_T1),
        (2, CATEGORY_LOW_CLEAR_T2),
        (3, CATEGORY_LOW_CLEAR_T3),
        (4, CATEGORY_LOW_CLEAR_T4),
        (5, CATEGORY_LOW_CLEAR_T5PLUS),
    ]

    selected = []

    for category in CATEGORIES:
        print(f"  [{category}]")
        cat = [e for e in entries if e["category"] == category]

        steam_pots = [e for e in cat if e["type"] == "Potential" and e["platform"] == "steam"]
        n = min(CATEGORY_POTENTIALS_STEAM, len(steam_pots))
        selected.extend(random.sample(steam_pots, n))
        if n < CATEGORY_POTENTIALS_STEAM:
            print(f"    WARNING: steam potentials: only {n}/{CATEGORY_POTENTIALS_STEAM} available")
        else:
            print(f"    {n} steam potentials")

        # Retro potentials not yet implemented; placeholder for when retro support is added
        retro_pots = [e for e in cat if e["type"] == "Potential" and e["platform"] == "retro"]
        n = min(CATEGORY_POTENTIALS_RETRO, len(retro_pots))
        selected.extend(random.sample(retro_pots, n))
        if retro_pots:
            print(f"    {n} retro potentials")

        for tier, count in LOW_CLEAR_TIERS:
            pool = [e for e in cat if e["type"] == "Low Clear" and e["lowClearTier"] == tier]
            n = min(count, len(pool))
            selected.extend(random.sample(pool, n))
            if n < count:
                print(f"    WARNING: T{tier} low clears: only {n}/{count} available")
            else:
                print(f"    {n} T{tier} low clears")

        for obj_type, count, label in [
            ("Primary", CATEGORY_UNCLEARED_PO, "PO"),
            ("Secondary", CATEGORY_UNCLEARED_SO, "SO"),
        ]:
            pool = [e for e in cat if e["type"] == "Uncleared" and e.get("objectiveType") == obj_type]
            n = min(count, len(pool))
            selected.extend(random.sample(pool, n))
            if n < count:
                print(f"    WARNING: uncleared {label}: only {n}/{count} available")
            else:
                print(f"    {n} uncleared {label}s")

    for entry in selected:
        entry["basePrice"] = None
        entry["currentPrice"] = None

        bp = -1
        match(entry['type']):
            case "Low Clear":
                match(entry['lowClearTier']):
                    case 1:
                        bp = BOUNTY_POINTS_LOW_CLEAR_T1
                    case 2:
                        bp = BOUNTY_POINTS_LOW_CLEAR_T2
                    case 3:
                        bp = BOUNTY_POINTS_LOW_CLEAR_T3
                    case 4:
                        bp = BOUNTY_POINTS_LOW_CLEAR_T4
                    case 5:
                        bp = BOUNTY_POINTS_LOW_CLEAR_T5PLUS
            case "Uncleared":
                bp = BOUNTY_POINTS_UNCLEARED_OBJECTIVE
            case "Potential":
                bp = BOUNTY_POINTS_POTENTIAL
        entry["bountyPoints"] = bp

    with open("selection.json", "w") as f:
        json.dump(selected, f, indent=2)

    pots = sum(1 for e in selected if e["type"] == "Potential")
    lcs = sum(1 for e in selected if e["type"] == "Low Clear")
    uncs = sum(1 for e in selected if e["type"] == "Uncleared")
    print(f"Done. Selected {len(selected)} entries: {pots} potentials, {lcs} low clears, {uncs} uncleareds. Stored in selection.json.")

def selectdata():
    """
    Reads selection.json, fetches Steam store prices for all steam entries,
    and writes back basePrice and currentPrice (in USD) to selection.json.
    Games with no purchasable listing (unlisted/delisted) are written to
    unlistedgames.json and flagged in selection.json.
    """
    import json
    import time
    import os

    with open("selection.json") as f:
        entries = json.load(f)

    steam_entries = [e for e in entries if e.get("platform") == "steam" and e.get("platformId")]
    print(f"Fetching Steam prices for {len(steam_entries)} entries...")

    BATCH = 100
    app_ids = list({e["platformId"] for e in steam_entries})
    price_map: dict[str, dict] = {}
    no_price_ids: list[str] = []

    for i in range(0, len(app_ids), BATCH):
        batch = app_ids[i:i + BATCH]
        joined = ",".join(batch)
        resp = requests.get(
            "https://store.steampowered.com/api/appdetails",
            params={"appids": joined, "filters": "price_overview", "cc": "us"},
        )
        resp.raise_for_status()
        data = resp.json()
        for app_id, result in data.items():
            if not result.get("success"):
                no_price_ids.append(app_id)
                continue
            app_data = result.get("data")
            if not isinstance(app_data, dict):
                no_price_ids.append(app_id)
                continue
            po = app_data.get("price_overview")
            if po:
                price_map[app_id] = {
                    "basePrice": f"${po['initial'] / 100:.2f}",
                    "currentPrice": f"${po['final'] / 100:.2f}",
                }
            # else: dict data but no price_overview = free game, not unlisted
        print(f"  Fetched batch {i // BATCH + 1}/{(len(app_ids) + BATCH - 1) // BATCH} ({len(price_map)} priced so far)...")
        if i + BATCH < len(app_ids):
            time.sleep(1)

    # Second pass: classify no-price entries as free or unlisted
    # Unlisted = success but package_groups is empty and not free
    unlisted_ids: set[str] = set()
    if no_price_ids:
        print(f"  Checking {len(no_price_ids)} no-price entries (free vs unlisted)...")
        for app_id in no_price_ids:
            resp = requests.get(
                "https://store.steampowered.com/api/appdetails",
                params={"appids": app_id, "cc": "us"},
            )
            resp.raise_for_status()
            result = resp.json().get(app_id, {})
            if not result.get("success"):
                unlisted_ids.add(app_id)
                continue
            app_data = result.get("data")
            if not isinstance(app_data, dict):
                unlisted_ids.add(app_id)
                continue
            if app_data.get("is_free"):
                print(f"    {app_data.get('name', app_id)}: free, skipping price")
            elif not app_data.get("package_groups"):
                print(f"    {app_data.get('name', app_id)}: unlisted")
                unlisted_ids.add(app_id)
            time.sleep(0.5)

    # Build a name lookup for unlisted entries
    pid_to_name = {e["platformId"]: e["gameName"] for e in steam_entries}

    # Load existing unlistedgames.json and merge
    unlisted_path = "unlistedgames.json"
    existing_unlisted: list[dict] = []
    if os.path.exists(unlisted_path):
        with open(unlisted_path) as f:
            existing_unlisted = json.load(f)
    existing_pids = {u["platformId"] for u in existing_unlisted}
    for pid in unlisted_ids:
        if pid not in existing_pids:
            existing_unlisted.append({"platformId": pid, "gameName": pid_to_name.get(pid, "Unknown")})
    with open(unlisted_path, "w") as f:
        json.dump(existing_unlisted, f, indent=2)
    if unlisted_ids:
        print(f"  {len(unlisted_ids)} unlisted game(s) added to unlistedgames.json. Re-run select to exclude them.")

    filled = 0
    for entry in entries:
        pid = entry.get("platformId")
        if entry.get("platform") != "steam" or not pid:
            continue
        if pid in price_map:
            entry["basePrice"] = price_map[pid]["basePrice"]
            entry["currentPrice"] = price_map[pid]["currentPrice"]
            filled += 1
        elif pid in unlisted_ids:
            entry["basePrice"] = "UNLISTED"
            entry["currentPrice"] = "UNLISTED"

    with open("selection.json", "w") as f:
        json.dump(entries, f, indent=2)

    print(f"Done. Filled prices for {filled}/{len(steam_entries)} steam entries. Saved to selection.json.")


def __pull_uncleareds() -> list[dict]:
    """
    Pulls all uncleared POs and SOs from the site and returns them in this object:
    {
        "type": "Primary" # or "Secondary"
        "gameCeId": "ee810c8f-d5c8-4a4b-9200-b89ea4a83901" # or any uuidv4
        "gameName": "Jupiter Hell"
        "objectiveCeId": "330648e3-d94e-47ae-b3e5-d5dc17385896" # or any uuidv4
        "objectiveName": "Eternal Doom (UNCLEARED)" # or any string
    }
    We can do this by going through /api/objectives using the ?orderBy=points metric.
    - limit is default 100 which is fine.
    - then we increase the offset
    - discard any COs
    - and stop once we find one with points > 0.
    We should also store this as uncleareds.json
    """
    import json

    BASE = "https://cedb.me/api"

    CLOWN_TOWN = '09f100aa-caa7-4154-a224-1c3e9277eea4'

    GENRE_TO_CATEGORY: dict[str, str] = {
        "4d43349a-43a8-4755-9d52-41ece63ec5b1": "Action",
        "ec499226-0913-4db1-890e-093b366bcb3c": "Arcade",
        "7f8676fe-4900-400b-9284-c073388d88f7": "Bullet Hell",
        "a6d00cc0-9481-47cb-bb52-a7011041915a": "First-Person",
        "3c3fd562-525c-4e24-a1fa-5b5eda85ebbd": "Platformer",
        "ffb558c1-5a45-4b8c-856c-e9622ce54f00": "Strategy",
    }

    print("Fetching game list from /api/games...")
    games_resp = requests.get(f"{BASE}/games")
    games_resp.raise_for_status()
    all_games = games_resp.json()
    game_cache: dict[str, dict] = {
        g["id"]: {
            "name": g["name"],
            "platform": g.get("platform"),
            "platformId": g.get("platformId"),
            "category": GENRE_TO_CATEGORY.get(g.get("genreId", "")),
        }
        for g in all_games
    }
    game_name_cache: dict[str, str] = {gid: info["name"] for gid, info in game_cache.items()}
    print(f"  Loaded {len(game_cache)} games.")

    with open("cedb_games.json", "w") as f:
        json.dump(game_cache, f, indent=2)
    print("  Saved cedb_games.json.")

    results = []
    offset = 0

    print("Fetching objectives...")
    while True:
        resp = requests.get(f"{BASE}/objectives", params={"sortBy": "points", "offset": offset})
        resp.raise_for_status()
        objectives = resp.json()

        if not objectives:
            break

        done = False
        for obj in objectives:
            if obj["points"] > 0:
                done = True
                break
            if obj["community"] or obj['type'] == 'community' or obj['gameId'] == CLOWN_TOWN:
                continue

            game_id = obj["gameId"]
            if game_id not in game_name_cache:
                print(f"  WARNING: gameId {game_id} not found in /api/games (objective: {obj['name']})")
                continue

            results.append({
                "type": obj["type"].capitalize(),
                "gameCeId": game_id,
                "gameName": game_name_cache[game_id],
                "category": game_cache[game_id].get("category"),
                "objectiveCeId": obj["id"],
                "objectiveName": obj["name"],
            })

        print(f"  Fetched offset {offset}-{offset + len(objectives) - 1}: {len(results)} uncleareds so far...")

        if done:
            break

        offset += 100

    with open("uncleareds.json", "w") as f:
        json.dump(results, f, indent=2)

    pos = sum(1 for r in results if r["type"] == "Primary")
    sos = sum(1 for r in results if r["type"] == "Secondary")
    unique_games = len({r["gameCeId"] for r in results})
    print(f"Done. Pulled {len(results)} uncleareds: {pos} POs and {sos} SOs from {unique_games} games. Stored in uncleareds.json.")

    return results


def totalinfo():
    """
    Prints a tally of totalinfo.json broken down by category and type:
    - Steam potentials
    - Low clears per tier (T1–T5+)
    - Uncleared POs
    - Uncleared SOs
    """
    import json

    with open("totalinfo.json") as f:
        entries = json.load(f)

    CATEGORIES = ["Action", "Arcade", "Bullet Hell", "First-Person", "Platformer", "Strategy"]
    TIERS = [1, 2, 3, 4, 5]

    for category in CATEGORIES:
        cat = [e for e in entries if e["category"] == category]
        steam_pots = sum(1 for e in cat if e["type"] == "Potential" and e["platform"] == "steam")
        lc = {t: sum(1 for e in cat if e["type"] == "Low Clear" and e["lowClearTier"] == t) for t in TIERS}
        pos = sum(1 for e in cat if e["type"] == "Uncleared" and e.get("objectiveType") == "Primary")
        sos = sum(1 for e in cat if e["type"] == "Uncleared" and e.get("objectiveType") == "Secondary")
        lc_str = "  ".join(f"T{t}:{lc[t]}" for t in TIERS)
        print(f"{category}: {steam_pots} potentials  {lc_str}  {pos} POs  {sos} SOs")


match(sys.argv[1]):
    case "pull":
        pull()
    case "pullce":
        __pull_uncleareds()
    case "pull+":
        pull()
        __pull_uncleareds()
    case "extract":
        extract()
    case "select":
        select()
    case "selectdata":
        selectdata()
    case "select+":
        select()
        selectdata()
    case "output":
        output()
    case "all":
        print("This will run the full pipeline in order:")
        print("  1. pull      — download potentials.xlsx and retro.xlsx from Google Sheets")
        print("  2. pullce    — fetch all uncleared POs/SOs and game data from cedb.me")
        print("  3. extract   — parse spreadsheets + CEDB data into totalinfo.json,")
        print("                 check every Steam game for unlisted status (makes Steam API calls)")
        print("  4. select    — randomly pick games per category from totalinfo.json → selection.json")
        print("                 (unlisted games in unlistedgames.json are excluded)")
        print("  5. selectdata— fetch current Steam prices for every selected game → updates selection.json")
        print("                 (any newly discovered unlisted games are added to unlistedgames.json)")
        print("  6. output    — format selection into Discord messages → output.json")
        print()
        confirm = input("Type YES to proceed: ").strip()
        if confirm != "YES":
            print("Aborted.")
        else:
            pull()
            __pull_uncleareds()
            extract()
            select()
            selectdata()
            output()
    case "totalinfo":
        totalinfo()
